from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from config import LOWER_HEADERS, RESULT_HEADERS


def get_sheet_excel(path_to_file: str) -> Worksheet | str:
    """Получение листа из Excel файла.

    Args:
        path_to_file (str): Полный путь до Excel файла.

    Returns:
        Worksheet | str: Объект Excel листа, иначе значение ошибка.
    """
    try:
        work_book = load_workbook(filename=path_to_file)
        return work_book.active
    except Exception as ex:
        return ex

def get_headers_and_row(sheet: Worksheet) -> tuple[list[Any], tuple[Cell]] | bool:
    """Получение наименований столбцов и их номер строки.

    Args:
        sheet (Worksheet): Лист в таблице Excel.

    Returns:
        tuple[list[Any], tuple[Cell]]: Наименования столбцов и номер строки, если не нашёл данные - False.
    """
    headers_counter = 0
    headers_table = []

    for row in sheet:
        for cell in row:
            if str(cell.value).lower() in LOWER_HEADERS:
                headers_counter += 1
            if headers_counter >= 2:
                headers_table = [_cell.value for _cell in row]
                return headers_table, row
    return False


def get_generated_data(sorted_data: list[list[Any]], headers_table: list[Any]) -> dict[str, list[Any]]:
    """Формирует словарь с результатирующими данными для записи в Excel.

    Args:
        sorted_data (list[list[Any]]): Отсортированный список списков.
        headers_table (list[Any]): Заголовки в Excel файле от пользователя.

    Returns:
        dict[str, list[Any]]: Словарь с результатирующими данными.
    """
    result_dict = {}

    for result_header in RESULT_HEADERS:
        for idx, header in enumerate(headers_table):
            if result_header.lower() == str(header).lower():
                for data in sorted_data:
                    if header not in result_dict:
                        result_dict[header] = []
                    result_dict[header].append(data[idx])
    return result_dict


def get_result_data(
        word_filter: str,
        sheet: Worksheet,
        headers_row: Cell,
        user_header: str,
        headers_table: list[Any]
        ) -> dict[str, list[Any]]:
    """Формирование и получение данных для записи в Excel файл.

    Args:
        word_filter (str): Значение, по которому идёт фильтрация.
        sheet (Worksheet): Лист в таблице Excel.
        headers_row (Cell): Строка, где находятся заголовки в таблице.
        user_header (str): Выбранный столбец, по которому будет фильтрация.
        headers_table (list[Any]): Наименования всех столбцов в Excel файле.

    Returns:
        dict[str, list[Any]]: Словарь с результатирующими данными.
    """
    for cell in headers_row:
        if str(cell.value).lower() == user_header.lower():
            column = cell.column_letter
            start_row = cell.row + 1
            sorted_data = []
            for number_row in range(start_row, sheet.max_row + 1):
                cell_value = sheet[f"{column}{number_row}"].value
                if str(cell_value).lower() == word_filter.lower():  # Игнорируем пустые ячейки
                    row_data = [cell.value for cell in sheet[number_row]]
                    sorted_data.append(row_data)
            return get_generated_data(sorted_data=sorted_data, headers_table=headers_table)


def get_result_excel(result_data: dict[str, list[Any]]) -> Workbook:
    """Получение объекта результата Excel данных.

    Args:
        result_data (dict[str, list[Any]]): Словарь с результатирующими данными.

    Returns:
        Workbook: Объект Excel данных.
    """
    result_workbook = Workbook()
    result_worksheet = result_workbook.active
    result_worksheet.title = "Отфильтрованные данные"
    result_worksheet.append(list(result_data.keys()))
    for row in zip(*result_data.values()):
        result_worksheet.append(row)

    return result_workbook
